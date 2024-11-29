import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import qrcode
import cv2
import sqlite3
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from tkcalendar import DateEntry 
from PIL import Image, ImageTk
import winsound
import openpyxl
from openpyxl.styles import Font
import tkinter.filedialog as fd
from tkinter import Tk, Button, Label, PhotoImage
from tkcalendar import Calendar

def init_db():
    conn = sqlite3.connect("attendance_system1.db")
    cursor = conn.cursor()
    # Employees table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            unique_id TEXT UNIQUE NOT NULL
        )
    ''')
    # Attendance table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unique_id TEXT NOT NULL,
            name TEXT NOT NULL,
            time_in TEXT,
            time_out TEXT
        )
    ''')
    conn.commit()
    conn.close()


# Add Employee and Generate QR Code
def add_employee():
    def submit_employee():
        name = name_entry.get()
        if not name.strip():
            messagebox.showerror("Error", "Employee name cannot be empty!")
            return

        conn = sqlite3.connect("attendance_system1.db")
        cursor = conn.cursor()

        unique_id = f"KAASA-{name}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        try:
            cursor.execute("INSERT INTO employees (name, unique_id) VALUES (?, ?)", (name, unique_id))
            conn.commit()
            qr = qrcode.make(unique_id)
            file_name = f"{name}_QR.png"
            qr.save(file_name)

            messagebox.showinfo("Success", f"Employee added successfully.\nQR Code saved as {file_name}.")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Employee already exists!")
        finally:
            conn.close()
            add_employee_window.destroy()

    # Create a new window for adding employee
    add_employee_window = tk.Toplevel()
    add_employee_window.title("Add Employee")
    add_employee_window.geometry("400x200")

    tk.Label(add_employee_window, text="Enter Employee Name:", font=("Helvetica", 12)).pack(pady=10)
    name_entry = tk.Entry(add_employee_window, font=("Helvetica", 12))
    name_entry.pack(pady=10)

    submit_button = tk.Button(add_employee_window, text="Add Employee", font=("Helvetica", 12), command=submit_employee)
    submit_button.pack(pady=10)

    cancel_button = tk.Button(add_employee_window, text="Cancel", font=("Helvetica", 12), command=add_employee_window.destroy)
    cancel_button.pack(pady=5)


# Scan QR Code and Record Attendance
def scan_qr():
    def display_details(unique_id, name, time_in, time_out=None):
        # Ensure that green_tick_box is available in the scope of display_details
        green_tick_box.pack(pady=20)
        uid_label.config(text=f"Unique ID: {unique_id}")
        name_label.config(text=f"Name: {name}")
        time_in_label.config(text=f"Time In: {time_in}")
        time_out_label.config(text=f"Time Out: {time_out if time_out else '-'}")

    # Create a new scanner window
    scanner_window = tk.Toplevel()
    scanner_window.title("QR Code Scanner")
    scanner_window.geometry("600x400")

    # Set background image
    # Keep reference to the image to prevent garbage collection

    # Text Labels on top of the background image
    tk.Label(scanner_window, text="QR Code Scanner", font=("Helvetica", 16), bg="grey").pack(pady=10)
    tk.Label(scanner_window, text="WELCOME TO KAASA", font=("Helvetica", 25), bg="cyan").pack(pady=10)
    tk.Label(scanner_window, text="The Future Of EV is here", font=("Helvetica", 25), bg="grey").pack(pady=10)

    # Green tick box frame
    green_tick_box = tk.Frame(scanner_window, bg="white", relief="solid", bd=1)
    green_tick_box.pack_forget()  # Hidden initially

    # Add green tick image
    tick_img = Image.open("1.png").resize((50, 50), Image.LANCZOS)
    tick_photo = ImageTk.PhotoImage(tick_img)
    tick_label = tk.Label(green_tick_box, image=tick_photo, bg="white")
    tick_label.image = tick_photo
    tick_label.pack(pady=10)

    # Information labels
    uid_label = tk.Label(green_tick_box, text="", font=("Helvetica", 12), bg="white", anchor="w")
    uid_label.pack(anchor="w", padx=20, pady=5)
    name_label = tk.Label(green_tick_box, text="", font=("Helvetica", 12), bg="white", anchor="w")
    name_label.pack(anchor="w", padx=20, pady=5)
    time_in_label = tk.Label(green_tick_box, text="", font=("Helvetica", 12), bg="white", anchor="w")
    time_in_label.pack(anchor="w", padx=20, pady=5)
    time_out_label = tk.Label(green_tick_box, text="", font=("Helvetica", 12), bg="white", anchor="w")
    time_out_label.pack(anchor="w", padx=20, pady=5)

    # Button to close the scanner
    def close_scanner():
        cap.release()
        cv2.destroyAllWindows()
        scanner_window.destroy()

    tk.Button(scanner_window, text="Close Scanner", command=close_scanner, font=("Helvetica", 12)).pack(pady=10)

    cap = cv2.VideoCapture(0)  # Initialize camera
    detector = cv2.QRCodeDetector()
    conn = sqlite3.connect("attendance_system1.db")
    cursor = conn.cursor()

    # Function to continuously scan QR codes
    def scan():
        ret, frame = cap.read()
        if ret:
            # Process the frame for QR codes
            data, bbox, _ = detector.detectAndDecode(frame)
            if data:
                unique_id = data.strip()
                cursor.execute("SELECT name FROM employees WHERE unique_id = ?", (unique_id,))
                result = cursor.fetchone()

                if result:
                    name = result[0]
                    today_date = datetime.now().strftime("%d-%m-%Y")
                    cursor.execute("SELECT * FROM attendance WHERE unique_id = ? AND time_in LIKE ?",
                                   (unique_id, f"{today_date}%"))
                    attendance_record = cursor.fetchone()

                    if attendance_record:
                        if attendance_record[4]:  # time_out already exists
                            display_details(unique_id, name, attendance_record[3], attendance_record[4])
                        else:
                            # Update time_out
                            now = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
                            cursor.execute("UPDATE attendance SET time_out = ? WHERE id = ?",
                                           (now, attendance_record[0]))
                            conn.commit()
                            display_details(unique_id, name, attendance_record[3], now)
                    else:
                        # Record time_in
                        now = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
                        cursor.execute("INSERT INTO attendance (unique_id, name, time_in) VALUES (?, ?, ?)",
                                       (unique_id, name, now))
                        conn.commit()
                        display_details(unique_id, name, now)

                    # Beep sound when QR is scanned
                    import winsound
                    winsound.Beep(1000, 500)  # Frequency, Duration in ms

                    # Delay to prevent repeated scans of the same code
                    scanner_window.after(3000, lambda: green_tick_box.pack_forget())
                # Short delay before next scan to avoid immediate repeated processing
                scanner_window.after(2000, scan)
            else:
                # Continue scanning in the next iteration
                scanner_window.after(10, scan)

    # Start scanning loop
    scan()

    # Handle graceful exit
    scanner_window.protocol("WM_DELETE_WINDOW", close_scanner)

def calculate_working_hours(time_in, time_out):
    time_format = "%d-%m-%Y %I:%M:%S %p"
    time_in_obj = datetime.strptime(time_in, time_format)
    time_out_obj = datetime.strptime(time_out, time_format)

    work_duration = time_out_obj - time_in_obj
    return work_duration
def view_working_hours():
    def load_attendance():
        conn = sqlite3.connect("attendance_system1.db")
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM employees")
        employees = cursor.fetchall()

        for row in tree.get_children():
            tree.delete(row)

        hourly_rate = 40  # Set the hourly rate

        for employee in employees:
            name = employee[1]
            unique_id = employee[2]

            cursor.execute("SELECT * FROM attendance WHERE unique_id = ?", (unique_id,))
            attendance_record = cursor.fetchall()

            monthly_hours = {}

            for record in attendance_record:
                time_in = record[3]
                time_out = record[4] if record[4] else 'N/A'
                if time_out != 'N/A':
                    # Calculate working hours
                    work_duration = calculate_working_hours(time_in, time_out)
                    month_year = datetime.strptime(time_in, "%d-%m-%Y %I:%M:%S %p").strftime("%B %Y")

                    if month_year not in monthly_hours:
                        monthly_hours[month_year] = timedelta()

                    monthly_hours[month_year] += work_duration

            # Now, add records for each month worked
            for month_year, total_hours in monthly_hours.items():
                total_seconds = total_hours.total_seconds()
                total_hours_decimal = total_seconds / 3600  # Convert to decimal hours
                salary = total_hours_decimal * hourly_rate
                tree.insert("", tk.END, values=(unique_id, name, month_year, str(total_hours), f"₹{salary:.2f}"))

        conn.close()

    def export_to_excel():
        # Get all data from the TreeView
        rows = tree.get_children()
        if not rows:
            tk.messagebox.showinfo("Export Failed", "No data to export!")
            return

        # Open file dialog for saving the Excel file
        file_path = fd.asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_path:
            return

        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Employee Working Hours"

        # Add headers
        headers = ["Unique ID", "Name", "Month", "Total Working Hours", "Salary"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Add data rows
        for row_num, item in enumerate(rows, start=2):
            values = tree.item(item, "values")
            for col_num, value in enumerate(values, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Save the Excel file
        workbook.save(file_path)
        tk.messagebox.showinfo("Export Successful", f"Data exported successfully to {file_path}")

    # Create a new window for viewing working hours with monthly totals
    working_hours_window = tk.Toplevel()
    working_hours_window.title("View Employee Working Hours (Monthly)")
    working_hours_window.geometry("800x400")

    tree = ttk.Treeview(working_hours_window, columns=("Unique ID", "Name", "Month", "Total Working Hours", "Salary"), show="headings")
    tree.heading("Unique ID", text="Unique ID")
    tree.heading("Name", text="Name")
    tree.heading("Month", text="Month")
    tree.heading("Total Working Hours", text="Total Working Hours")
    tree.heading("Salary", text="Salary")
    tree.pack(fill=tk.BOTH, expand=True)

    # Add Load Attendance button
    load_button = tk.Button(working_hours_window, text="Load Attendance", command=load_attendance)
    load_button.pack(pady=10)

    # Add Export to Excel button
    export_button = tk.Button(working_hours_window, text="Export to Excel", command=export_to_excel)
    export_button.pack(pady=5)

    # Add Close button
    close_button = tk.Button(working_hours_window, text="Close", command=working_hours_window.destroy)
    close_button.pack(pady=5)

# View Attendance History
def view_attendance():
    def load_records():
        conn = sqlite3.connect("attendance_system1.db")
        cursor = conn.cursor()

        selected_date = date_filter.get()
        query = "SELECT * FROM attendance"
        if selected_date:
            query += f" WHERE time_in LIKE '{selected_date}%'"

        cursor.execute(query)
        records = cursor.fetchall()
        conn.close()

        for row in tree.get_children():
            tree.delete(row)

        for record in records:
            time_in = record[3] if record[3] else "-"
            time_out = record[4] if record[4] else "-"
            working_hours = calculate_working_hours(record[3], record[4]) if record[3] and record[4] else "-"
            tree.insert("", "end", values=(record[0], record[1], record[2], time_in, time_out, working_hours))

    def calculate_working_hours(time_in, time_out):
        time_in = datetime.strptime(time_in, "%d-%m-%Y %I:%M:%S %p")
        time_out = datetime.strptime(time_out, "%d-%m-%Y %I:%M:%S %p")
        delta = time_out - time_in
        return str(delta)

    def add_attendance():
        def save_attendance():
            unique_id = entry_unique_id.get()
            name = entry_name.get()
            time_in = entry_time_in.get()
            time_out = entry_time_out.get()
            
            if unique_id and name and time_in and time_out:
                conn = sqlite3.connect("attendance_system1.db")
                cursor = conn.cursor()
                
                query = "INSERT INTO attendance (unique_id, name, time_in, time_out) VALUES (?, ?, ?, ?)"
                cursor.execute(query, (unique_id, name, time_in, time_out))
                conn.commit()
                conn.close()
                load_records()  # Reload the records to display the updated list
                add_window.destroy()  # Close the "Add Attendance" window
            else:
                tk.messagebox.showerror("Input Error", "All fields are required!")

        add_window = tk.Toplevel()
        add_window.title("Add Attendance")
        add_window.geometry("400x300")

        tk.Label(add_window, text="Add Attendance Record", font=("Helvetica", 16)).pack(pady=10)

        tk.Label(add_window, text="Unique ID:", font=("Helvetica", 12)).pack(pady=5)
        entry_unique_id = tk.Entry(add_window, font=("Helvetica", 12))
        entry_unique_id.pack(pady=5)

        tk.Label(add_window, text="Name:", font=("Helvetica", 12)).pack(pady=5)
        entry_name = tk.Entry(add_window, font=("Helvetica", 12))
        entry_name.pack(pady=5)

        tk.Label(add_window, text="Time In (DD-MM-YYYY hh:mm:ss AM/PM):", font=("Helvetica", 12)).pack(pady=5)
        entry_time_in = tk.Entry(add_window, font=("Helvetica", 12))
        entry_time_in.pack(pady=5)

        tk.Label(add_window, text="Time Out (DD-MM-YYYY hh:mm:ss AM/PM):", font=("Helvetica", 12)).pack(pady=5)
        entry_time_out = tk.Entry(add_window, font=("Helvetica", 12))
        entry_time_out.pack(pady=5)

        tk.Button(add_window, text="Save", font=("Helvetica", 12), command=save_attendance).pack(pady=10)

    # Attendance Window
    view_window = tk.Toplevel()
    view_window.title("View Attendance")
    view_window.geometry("800x500")

    tk.Label(view_window, text="View Attendance Records", font=("Helvetica", 16)).pack(pady=10)

    frame = tk.Frame(view_window)
    frame.pack(pady=10)

    tk.Label(frame, text="Select Date (DD-MM-YYYY):", font=("Helvetica", 12)).grid(row=0, column=0, padx=5)
    date_filter = DateEntry(frame, date_pattern="dd-MM-yyyy", font=("Helvetica", 12))  # Calendar widget
    date_filter.grid(row=0, column=1, padx=5)
    tk.Button(frame, text="Apply Filter", font=("Helvetica", 12), command=load_records).grid(row=0, column=2, padx=5)

    tree = ttk.Treeview(view_window, columns=("ID", "Unique ID", "Name", "Time In", "Time Out", "Working Hours"), show="headings")
    tree.pack(pady=20, fill="both", expand=True)

    tree.heading("ID", text="ID")
    tree.heading("Unique ID", text="Unique ID")
    tree.heading("Name", text="Name")
    tree.heading("Time In", text="Time In")
    tree.heading("Time Out", text="Time Out")
    tree.heading("Working Hours", text="Working Hours")

    tree.column("ID", width=50)
    tree.column("Unique ID", width=150)
    tree.column("Name", width=150)
    tree.column("Time In", width=100)
    tree.column("Time Out", width=100)
    tree.column("Working Hours", width=120)

    load_records()

    # Add Attendance Button
    tk.Button(view_window, text="Add Attendance", font=("Helvetica", 12), command=add_attendance).pack(pady=10)

    # Close Button
    tk.Button(view_window, text="Close", font=("Helvetica", 12), command=view_window.destroy).pack(pady=10)


# Export to PDF
def export_to_pdf():
    conn = sqlite3.connect("attendance_system1.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM employees")
    employees = cursor.fetchall()

    # Create a PDF file to store the attendance report
    pdf_filename = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
    if not pdf_filename:
        return

    pdf = canvas.Canvas(pdf_filename, pagesize=letter)
    pdf.setFont("Helvetica", 12)

    pdf.drawString(100, 750, "Employee Attendance Report")
    pdf.drawString(100, 735, f"Generated on: {datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')}")

    y_position = 700

    for employee in employees:
        name = employee[1]
        unique_id = employee[2]
        cursor.execute("SELECT * FROM attendance WHERE unique_id = ?", (unique_id,))
        attendance_record = cursor.fetchall()

        for record in attendance_record:
            pdf.drawString(100, y_position, f"Unique ID: {unique_id}")
            pdf.drawString(200, y_position, f"Name: {name}")
            pdf.drawString(300, y_position, f"Time In: {record[3]}")
            pdf.drawString(400, y_position, f"Time Out: {record[4] if record[4] else 'N/A'}")
            y_position -= 20
            if y_position < 100:
                pdf.showPage()
                y_position = 750

    pdf.save()
    messagebox.showinfo("Export to PDF", "Attendance exported to PDF successfully.")

# Delete Record
def delete_record():
    def confirm_delete():
        record_id = delete_entry.get()
        if not record_id.strip():
            messagebox.showerror("Error", "Please provide a valid record ID.")
            return

        conn = sqlite3.connect("attendance_system1.db")
        cursor = conn.cursor()
        cursor.execute("DELETE FROM attendance WHERE id = ?", (record_id,))
        conn.commit()

        messagebox.showinfo("Success", "Attendance record deleted successfully.")
        delete_window.destroy()

    delete_window = tk.Toplevel()
    delete_window.title("Delete Attendance Record")
    delete_window.geometry("400x200")

    tk.Label(delete_window, text="Enter Attendance Record ID to delete:", font=("Helvetica", 12)).pack(pady=10)
    delete_entry = tk.Entry(delete_window, font=("Helvetica", 12))
    delete_entry.pack(pady=10)

    tk.Button(delete_window, text="Delete Record", command=confirm_delete, font=("Helvetica", 12)).pack(pady=10)
    tk.Button(delete_window, text="Cancel", command=delete_window.destroy, font=("Helvetica", 12)).pack(pady=5)

def edit_attendance(tree):
    # Get the selected row
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showerror("Error", "Please select an entry to edit.")
        return

    # Fetch the data of the selected row
    row_data = tree.item(selected_item)["values"]
    if not row_data:
        messagebox.showerror("Error", "Unable to fetch the selected data.")
        return

    # Create a pop-up window for editing
    edit_window = tk.Toplevel()
    edit_window.title("Edit Attendance")
    edit_window.geometry("400x300")

    # Labels and entry fields for each column
    tk.Label(edit_window, text="ID:").pack(pady=5)
    id_entry = tk.Entry(edit_window)
    id_entry.pack(pady=5)
    id_entry.insert(0, row_data[0])
    id_entry.configure(state="readonly")

    tk.Label(edit_window, text="Name:").pack(pady=5)
    name_entry = tk.Entry(edit_window)
    name_entry.pack(pady=5)
    name_entry.insert(0, row_data[1])

    tk.Label(edit_window, text="Date:").pack(pady=5)
    date_entry = tk.Entry(edit_window)
    date_entry.pack(pady=5)
    date_entry.insert(0, row_data[2])

    tk.Label(edit_window, text="Time In:").pack(pady=5)
    time_in_entry = tk.Entry(edit_window)
    time_in_entry.pack(pady=5)
    time_in_entry.insert(0, row_data[3])

    tk.Label(edit_window, text="Time Out:").pack(pady=5)
    time_out_entry = tk.Entry(edit_window)
    time_out_entry.pack(pady=5)
    time_out_entry.insert(0, row_data[4])

    # Save changes button
    def save_changes():
        new_name = name_entry.get()
        new_date = date_entry.get()
        new_time_in = time_in_entry.get()
        new_time_out = time_out_entry.get()

        # Update the selected row in TreeView
        tree.item(selected_item, values=(row_data[0], new_name, new_date, new_time_in, new_time_out))

        # Save changes to the file or database
        save_attendance_changes(row_data[0], new_name, new_date, new_time_in, new_time_out)
        
        edit_window.destroy()
        messagebox.showinfo("Success", "Attendance updated successfully!")

    save_button = tk.Button(edit_window, text="Save Changes", command=save_changes)
    save_button.pack(pady=10)

def save_attendance_changes(id, name, date, time_in, time_out):
    # Logic to save updated data to file or database
    # For example, update the CSV file:
    updated_data = []
    with open("attendance.csv", "r") as file:
        reader = csv.reader(file)
        for row in reader:
            if row[0] == str(id):
                updated_data.append([id, name, date, time_in, time_out])
            else:
                updated_data.append(row)

    with open("attendance.csv", "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(updated_data)
def view_employee():
    # Create a new window to display employee details
    view_window = tk.Tk()
    view_window.title("Employee Details")
    view_window.geometry("800x600")
    view_window.config(bg="#f0f0f0")

    # Title for the employee details window
    title_label = tk.Label(
        view_window,
        text="Employee Details",
        font=("Helvetica", 25, "bold"),
        fg="white",
        bg="black",
        relief="solid",
        width=25,
        height=2
    )
    title_label.pack(pady=20)

    # Treeview frame
    tree_frame = tk.Frame(view_window, bg="#f0f0f0")
    tree_frame.pack(pady=10, padx=20, fill="both", expand=True)

    # Scrollbar for the treeview
    tree_scroll = tk.Scrollbar(tree_frame)
    tree_scroll.pack(side="right", fill="y")

    # Treeview widget
    tree = ttk.Treeview(
        tree_frame,
        yscrollcommand=tree_scroll.set,
        columns=("Unique ID", "Name"),
        show="headings",
        height=20
    )
    tree.pack(side="left", fill="both", expand=True)
    tree_scroll.config(command=tree.yview)

    # Define columns
    tree.heading("Unique ID", text="Unique ID", anchor="w")
    tree.heading("Name", text="Name", anchor="w")
    tree.column("Unique ID", anchor="w", width=200)
    tree.column("Name", anchor="w", width=400)

    # Database connection
    conn = sqlite3.connect("attendance_system1.db")
    cursor = conn.cursor()

    # Fetch all employees from the database
    cursor.execute("SELECT unique_id, name FROM employees")
    employees = cursor.fetchall()

    # Insert data into the treeview
    for employee in employees:
        tree.insert("", "end", values=employee)

    # Close button
    close_button = tk.Button(
        view_window,
        text="Close",
        command=view_window.destroy,
        font=("Helvetica", 14),
        bg="#dc3545",  # Red color
        fg="white",
        relief="raised",
        width=13,
        height=2
    )
    close_button.pack(pady=20)

    view_window.mainloop()
def view_employee_attendance():
    def search_employee():
        # Get the entered employee name
        search_name = name_entry.get().strip()

        # Clear the previous result
        for widget in result_frame.winfo_children():
            widget.destroy()

        if not search_name:
            tk.Label(result_frame, text="Please enter a valid name.", font=("Helvetica", 14), fg="red").pack()
            return

        # Fetch employee data from the database
        cursor.execute("SELECT unique_id, name FROM employees WHERE name LIKE ?", (f"%{search_name}%",))
        employee = cursor.fetchone()

        if not employee:
            tk.Label(result_frame, text="Employee not found.", font=("Helvetica", 14), fg="red").pack()
            return

        unique_id, name = employee

        # Fetch attendance data for the employee
        cursor.execute(
            """
            SELECT time_in, time_out FROM attendance WHERE unique_id = ?
            """,
            (unique_id,),
        )
        attendance_records = cursor.fetchall()

        if not attendance_records:
            tk.Label(result_frame, text="No attendance records found for this employee.", font=("Helvetica", 14)).pack()
            return

        # Calculate monthly working hours
        monthly_data = {}
        for time_in, time_out in attendance_records:
            if time_out:
                time_in_dt = datetime.strptime(time_in, "%d-%m-%Y %I:%M:%S %p")
                time_out_dt = datetime.strptime(time_out, "%d-%m-%Y %I:%M:%S %p")
                working_hours = (time_out_dt - time_in_dt).total_seconds() / 3600

                # Group data by month
                month = time_in_dt.strftime("%B %Y")
                if month not in monthly_data:
                    monthly_data[month] = {"total_hours": 0, "dates": []}

                monthly_data[month]["total_hours"] += working_hours
                monthly_data[month]["dates"].append(
                    {"date": time_in_dt.strftime("%d-%m-%Y"), "working_hours": working_hours}
                )

        # Display the table
        tk.Label(result_frame, text=f"Attendance Report for {name}", font=("Helvetica", 16, "bold")).pack(pady=10)

        tree = ttk.Treeview(result_frame, columns=("Month", "Total Hours", "Details"), show="headings")
        tree.heading("Month", text="Month")
        tree.heading("Total Hours", text="Total Hours")
        tree.heading("Details", text="Details")

        tree.column("Month", width=200, anchor="center")
        tree.column("Total Hours", width=150, anchor="center")
        tree.column("Details", width=150, anchor="center")

        # Populate the table with monthly data
        for month, data in monthly_data.items():
            tree.insert("", "end", values=(month, f"{data['total_hours']:.2f}", "View Details"))

        tree.pack(pady=10, fill="both", expand=True)

        # Handle row selection to view day-wise details
        def on_row_select(event):
            selected_item = tree.selection()
            if selected_item:
                item = tree.item(selected_item)
                selected_month = item["values"][0]

                # Display day-wise details for the selected month
                display_monthly_details(selected_month, monthly_data[selected_month]["dates"])

        tree.bind("<<TreeviewSelect>>", on_row_select)

    def display_monthly_details(month, daily_data):
        # Open a new window for day-wise details
        details_window = tk.Toplevel(attendance_window)
        details_window.title(f"Day-wise Attendance for {month}")
        details_window.geometry("600x400")
        details_window.config(bg="#f0f0f0")

        tk.Label(
            details_window,
            text=f"Day-wise Attendance for {month}",
            font=("Helvetica", 16, "bold"),
            fg="#333",
        ).pack(pady=10)

        # Create a table for day-wise details
        daily_tree = ttk.Treeview(details_window, columns=("Date", "Working Hours"), show="headings")
        daily_tree.heading("Date", text="Date")
        daily_tree.heading("Working Hours", text="Working Hours")

        daily_tree.column("Date", width=200, anchor="center")
        daily_tree.column("Working Hours", width=150, anchor="center")

        # Populate the table with day-wise data
        for record in daily_data:
            daily_tree.insert("", "end", values=(record["date"], f"{record['working_hours']:.2f} hrs"))

        daily_tree.pack(pady=10, fill="both", expand=True)

        # Close button
        close_button = tk.Button(
            details_window,
            text="Close",
            command=details_window.destroy,
            font=("Helvetica", 14),
            bg="#dc3545",
            fg="white",
        )
        close_button.pack(pady=10)

    # Create the main attendance window
    attendance_window = tk.Tk()
    attendance_window.title("Employee Attendance Report")
    attendance_window.geometry("800x600")
    attendance_window.config(bg="#f0f0f0")

    # Title
    title_label = tk.Label(
        attendance_window,
        text="Employee Attendance Report",
        font=("Helvetica", 25, "bold"),
        fg="white",
        bg="cyan",
        relief="solid",
        width=30,
        height=2,
    )
    title_label.pack(pady=20)

    # Search frame
    search_frame = tk.Frame(attendance_window, bg="#f0f0f0")
    search_frame.pack(pady=10)

    tk.Label(search_frame, text="Enter Employee Name:", font=("Helvetica", 14), bg="#f0f0f0").pack(side="left", padx=5)
    name_entry = tk.Entry(search_frame, font=("Helvetica", 14), width=30)
    name_entry.pack(side="left", padx=5)

    search_button = tk.Button(
        search_frame,
        text="Search",
        command=search_employee,
        font=("Helvetica", 14),
        bg="#28a745",
        fg="white",
        relief="raised",
    )
    search_button.pack(side="left", padx=5)

    # Result frame
    result_frame = tk.Frame(attendance_window, bg="#f0f0f0")
    result_frame.pack(pady=20, fill="both", expand=True)

    # Database connection
    conn = sqlite3.connect("attendance_system1.db")
    cursor = conn.cursor()

    # Close button
    close_button = tk.Button(
        attendance_window,
        text="Close",
        command=attendance_window.destroy,
        font=("Helvetica", 14),
        bg="#dc3545",
        fg="white",
        relief="raised",
    )
    close_button.pack(pady=10)

    attendance_window.mainloop()

   
    
def login():
    def validate_login():
        username = username_entry.get()
        password = password_entry.get()

        # Simple login validation (you can replace this with a proper authentication system)
        if username == "admin" and password == "1234":
            messagebox.showinfo("Login Success", "Welcome to KAASA Attendance System!")
            login_window.destroy()  # Close the login window
            main_window()  # Open the main window (KAASA Attendance System)
        else:
            messagebox.showerror("Login Error", "Invalid credentials, please try again.")

    # Create the login window
    login_window = tk.Tk()
    login_window.title("KAASA ADMIN LOGIN")
    login_window.geometry("1920x1080")

    # Load and set the background image
    background_image = Image.open("3.jpg")  # Ensure you have a 'background.jpg' image
    background_image = background_image.resize((1500, 800), Image.LANCZOS)
    background_photo = ImageTk.PhotoImage(background_image)

    # Set the background image in the window
    background_label = tk.Label(login_window, image=background_photo)
    background_label.place(relwidth=1, relheight=1)

    # Create a frame for the login form to make it look neat
    login_frame = tk.Frame(login_window, bg="grey", bd=5)
    login_frame.place(relx=0.5, rely=0.5, relwidth=0.8, relheight=0.5, anchor="center")

    # Add an admin icon above the login title
    admin_icon = Image.open("2.jpg")  # Ensure you have an 'admin_icon.png' file
    admin_icon = admin_icon.resize((80, 80), Image.LANCZOS)
    admin_icon_photo = ImageTk.PhotoImage(admin_icon)

    admin_icon_label = tk.Label(login_frame, image=admin_icon_photo, bg="white")
    admin_icon_label.pack(pady=10)

    # Add title label
    title_label = tk.Label(login_frame, text="KAASA ADMIN LOGIN", font=("Helvetica", 18, "bold"), bg="white")
    title_label.pack(pady=5)

    # Username and password labels and entries
    tk.Label(login_frame, text="Username:", font=("Helvetica", 12), bg="white").pack(pady=5)
    username_entry = tk.Entry(login_frame, font=("Helvetica", 12))
    username_entry.pack(pady=5)

    tk.Label(login_frame, text="Password:", font=("Helvetica", 12), bg="white").pack(pady=5)
    password_entry = tk.Entry(login_frame, font=("Helvetica", 12), show="*")
    password_entry.pack(pady=5)

    # Login button
    login_button = tk.Button(login_frame, text="Login", font=("Helvetica", 14), bg="#4CAF50", fg="white", command=validate_login)
    login_button.pack(pady=10)

    # Run the login window
    login_window.mainloop()

# Initialize database and show the login window


# Modify main function to first show login window
# Logout function
def logout():
    response = messagebox.askyesno("Logout", "Are you sure you want to log out?")
    if response:
        # Close the current main window
        main_window_window.destroy()
        
# Modify main window function to include logout button

def main_window():
    global main_window_window  # Declare global reference to the main window

    # Create the main KAASA Attendance System window
    main_window_window = tk.Tk()
    main_window_window.title("KAASA Attendance System")
    main_window_window.geometry("1500x800")

    # Label for title
    tk.Label(main_window_window, text="Master Setting Dashboard System", font=("Helvetica", 20, "bold")).pack(pady=20)

    # Create a frame to hold the buttons
    button_frame = tk.Frame(main_window_window)  # Use tk.Frame instead of Frame
    button_frame.pack(pady=20)  # Add some padding for spacing

    # (Rest of the code remains unchanged...)

    # Function to resize an image
    def resize_image(image_path, size):
        img = Image.open(image_path)
        img = img.resize(size, Image.Resampling.LANCZOS)  # Use LANCZOS for high-quality resizing
        return ImageTk.PhotoImage(img)

    # Load and resize images for buttons
    add_employee_img = resize_image("C:/KAASA/icons/addemployee.png", (64, 64))  # Adjust size as needed
    view_attendance_img = resize_image("C:/KAASA/icons/view attendance.png", (64, 64))
    view_working_hours_img = resize_image("C:/KAASA/icons/viewworkinghours.png", (64, 64))
    export_pdf_img = resize_image("C:/KAASA/icons/export.png", (64, 64))
    delete_record_img = resize_image("C:/KAASA/icons/delete.png", (64, 64))
    logout_img = resize_image("C:/KAASA/icons/logout.png", (64, 64))
    view_employee_img=resize_image("C:/KAASA/icons/view employee.png", (64, 64))
    view_employee_list=resize_image("C:/KAASA/icons/attendance.png", (64, 64))
    offline_attendance=resize_image("C:/KAASA/icons/offline attendance.png", (64, 64))

    # Buttons inside the frame
    Button(
        button_frame,
        text="Add Employee",
        font=("Helvetica", 12),
        image=add_employee_img,
        compound="top",
        command=add_employee  # Link to your existing function
    ).grid(row=0, column=0, padx=20, pady=10)

    Button(
        button_frame,
        text="View Attendance",
        font=("Helvetica", 12),
        image=view_attendance_img,
        compound="top",
        command=view_attendance  # Link to your existing function
    ).grid(row=0, column=1, padx=20, pady=10)

    Button(
        button_frame,
        text="View Working Hours",
        font=("Helvetica", 12),
        image=view_working_hours_img,
        compound="top",
        command=view_working_hours  # Link to your existing function
    ).grid(row=0, column=2, padx=20, pady=10)

    Button(
        button_frame,
        text="Export to PDF",
        font=("Helvetica", 12),
        image=export_pdf_img,
        compound="top",
        command=export_to_pdf  # Link to your existing function
    ).grid(row=1, column=0, padx=20, pady=10)

    Button(
        button_frame,
        text="Delete Record",
        font=("Helvetica", 12),
        image=delete_record_img,
        compound="top",
        command=delete_record  # Link to your existing function
    ).grid(row=1, column=1, padx=20, pady=10)

    Button(
        button_frame,
        text="Logout",
        font=("Helvetica", 12),
        image=logout_img,
        compound="top",
        command=logout  # Link to your existing function
    ).grid(row=1, column=2, padx=20, pady=20)
    
    Button(
        button_frame,
        text="View Employees",
        font=("Helvetica", 12),
        image=view_employee_img,
        compound="top",
        command=view_employee  # Link to your existing function
    ).grid(row=2, column=0, padx=20, pady=10)
    Button(
        button_frame,
        text="Attendance list Month wise",
        font=("Helvetica", 12),
        image=view_employee_list,
        compound="top",
        command=view_employee_attendance  # Link to your existing function
    ).grid(row=2, column=1, padx=20, pady=10)
    Button(
        button_frame,
        text="manual Present",
        font=("Helvetica", 12),
        image=view_employee_list,
        compound="top",
        command=mark_all_present,
    ).grid(row=2, column=2, padx=20, pady=10)


    # Keep references to images to prevent garbage collection
    main_window_window.add_employee_img = add_employee_img
    main_window_window.view_attendance_img = view_attendance_img
    main_window_window.view_working_hours_img = view_working_hours_img
    main_window_window.export_pdf_img = export_pdf_img
    main_window_window.delete_record_img = delete_record_img
    main_window_window.logout_img = logout_img
    # Start the main event loop
    main_window_window.mainloop()
# Initialize the database and login window
init_db()
login()  # First, show the login window
 # Call the login function before opening the main system window


